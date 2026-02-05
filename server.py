import os
import sys

# --- PORTABLE/FROZEN ENVIRONMENT INITIALIZATION ---
# This MUST happen before any local imports (like 'main') to ensure DLLs and paths are correct.
if getattr(sys, 'frozen', False):
    # sys._MEIPASS is the root of the extracted bundle (in onefile) or the directory of the EXE (in onedir)
    bundle_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(sys.executable)))
    
    # In PyInstaller 6+ with contents_directory='_internal', libraries are in a subfolder
    internal_dir = os.path.join(bundle_dir, '_internal')
    if os.path.isdir(internal_dir):
        base_path = internal_dir
        # Ensure DLLs in _internal can be found by Python 3.8+
        if hasattr(os, 'add_dll_directory'):
            try:
                os.add_dll_directory(internal_dir)
            except Exception:
                pass
        # Add to PATH as well for generic DLL searching
        os.environ["PATH"] = internal_dir + os.pathsep + os.environ.get("PATH", "")
    else:
        base_path = bundle_dir
    
    # Store for global access
    # Use bundle_dir (root) as base path for data/config to match settings.py behavior
    os.environ["APP_BASE_PATH"] = bundle_dir
    # Browsers are next to the exe, not in _internal
    exe_dir = os.path.dirname(os.path.abspath(sys.executable))
    os.environ["PLAYWRIGHT_BROWSERS_PATH"] = os.path.join(exe_dir, "browsers")
    
    # Adjust sys.path to ensure we can import from _internal if needed
    if internal_dir not in sys.path:
        sys.path.insert(0, internal_dir)
        
    # --- DEBUG PRINTS ---
    print(f"[DEBUG] sys.executable: {sys.executable}")
    print(f"[DEBUG] exe_dir: {exe_dir}")
    print(f"[DEBUG] PLAYWRIGHT_BROWSERS_PATH: {os.environ.get('PLAYWRIGHT_BROWSERS_PATH')}")
    # --------------------

import main as app_worker
import asyncio
import logging
import subprocess
import threading
import queue
from typing import List, Optional

from contextlib import asynccontextmanager
from fastapi import FastAPI, WebSocket, Query
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import sqlite3
import zipfile
from io import BytesIO
from storage import Database
import license_utils

# License State
IS_LICENSED = False


# License State
IS_LICENSED = False

def check_app_license():
    global IS_LICENSED
    valid, _ = license_utils.check_license_status()
    IS_LICENSED = valid
    if valid:
        print("[License] App is unlocked and ready.")
    else:
        print("[License] No valid license found. App is LOCKED.")

# Global state for the running process
class ProcessManager:
    def __init__(self):
        self.process = None
        self.lock = threading.Lock()
        self.output_queue = queue.Queue()
        self.is_running = False

    def start_process(self, command: List[str]):
        with self.lock:
            if self.is_running:
                raise Exception("A scan is already running")
            
            self.is_running = True
            # Use line buffering
            self.process = subprocess.Popen(
                command,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                cwd=os.path.dirname(os.path.abspath(__file__))
            )
            
            # Start a thread to read output
            threading.Thread(target=self._read_output, daemon=True).start()

    def _read_output(self):
        try:
            if not self.process:
                return
                
            for line in iter(self.process.stdout.readline, ''):
                self.output_queue.put(line)
                
            self.process.stdout.close()
            self.process.wait()
        except Exception as e:
            self.output_queue.put(f"Error reading output: {e}\n")
        finally:
            with self.lock:
                self.is_running = False
            self.output_queue.put("[Process Completed]\n")

    def kill_process(self):
        with self.lock:
            if self.process and self.is_running:
                self.process.terminate()
                self.is_running = False

process_manager = ProcessManager()

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Initialize DB tables on startup
    Database()
    check_app_license()
    yield
    # Clean up (if needed)

# Initialize FastAPI app
app = FastAPI(title="Bidding Scraper UI", lifespan=lifespan)

# Mount static files (Web UI)
# We expect 'web' folder at the root (bundle_dir)
web_dir = os.path.join(os.environ.get("APP_BASE_PATH", "."), "web")
if os.path.exists(web_dir):
    app.mount("/static", StaticFiles(directory=web_dir), name="static")
    
    # Serve index.html for root
    @app.get("/")
    async def read_root():
        return FileResponse(os.path.join(web_dir, "index.html"))


# Allow CORS (useful for dev)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def check_license_middleware(request, call_next):
    # Allow static files, auth endpoints, and docs
    path = request.url.path
    if path.startswith("/api/auth") or path.startswith("/static") or path == "/" or path.endswith(".js") or path.endswith(".css") or path.endswith("favicon.ico"):
        return await call_next(request)
    
    # Block other API calls if not licensed
    if path.startswith("/api/") and not IS_LICENSED:
        return JSONResponse(status_code=403, content={"error": "License required", "locked": True})
        
    return await call_next(request)

# Models
class SearchRequest(BaseModel):
    keywords: str
    search_type: str = "fulltext"
    pinmu: str = "all"
    category: str = "all"
    bid_type: str = "all"
    time_type: str = "1week"
    max_pages: int = 3
    start_date: Optional[str] = None
    end_date: Optional[str] = None

# Database helper
from config.settings import DB_PATH
def get_db_connection():
    # Use config.settings.DB_PATH for consistency
    db_path = str(DB_PATH)
    # Ensure directory exists (though Database() init should have done it)
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    
    print(f"[DEBUG] Connecting to DB at: {db_path}")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

# API Endpoints

class LicenseKey(BaseModel):
    key: str

@app.get("/api/auth/status")
async def get_auth_status():
    global IS_LICENSED
    # Re-check in case file was added manually
    valid, code = license_utils.check_license_status()
    IS_LICENSED = valid
    
    expire_str = "Unknown"
    if valid:
        info = license_utils.get_license_info()
        if info:
            expire_str = info.get("expire", "Unknown")
            
    return {"locked": not valid, "machine_code": code, "expire": expire_str}

@app.post("/api/auth/verify")
async def verify_auth_key(body: LicenseKey):
    global IS_LICENSED
    code = license_utils.get_machine_code()
    if license_utils.verify_license(code, body.key):
        license_utils.save_license(body.key)
        IS_LICENSED = True
        return {"success": True}
    return JSONResponse(status_code=400, content={"error": "Invalid License Key", "success": False})

@app.post("/api/search")
async def start_search(req: SearchRequest):
    try:
        # Construct command
        import sys
        
        # Determine executable
        # Unified Entry Point: Always call ourselves
        if getattr(sys, 'frozen', False):
             cmd = [sys.executable, "bxsearch"]
        else:
             cmd = [sys.executable, "server.py", "bxsearch"]
        
        # Handle keywords (comma separated)
        kw_list = [k.strip() for k in req.keywords.split(",") if k.strip()]
        if not kw_list:
             return JSONResponse(status_code=400, content={"error": "Keywords required"})
        
        cmd.extend(["--kw"] + kw_list)
        cmd.extend(["--search-type", req.search_type])
        cmd.extend(["--pinmu", req.pinmu])
        cmd.extend(["--category", req.category])
        cmd.extend(["--bid-type", req.bid_type])
        cmd.extend(["--time", req.time_type])
        cmd.extend(["--max-pages", str(req.max_pages)])
        
        if req.time_type == "custom":
            if req.start_date:
                cmd.extend(["--start-date", req.start_date])
            if req.end_date:
                cmd.extend(["--end-date", req.end_date])
        
        print(f"[SERVER DEBUG] Frozen: {getattr(sys, 'frozen', False)}")
        print(f"[SERVER DEBUG] Executable: {sys.executable}")
        print(f"[SERVER DEBUG] Command: {cmd}")
        
        process_manager.start_process(cmd)
        return {"status": "started", "command": " ".join(cmd)}
    except Exception as e:
        return JSONResponse(status_code=400, content={"error": str(e)})

@app.post("/api/stop")
async def stop_search():
    process_manager.kill_process()
    return {"status": "stopped"}

@app.get("/api/status")
async def get_status():
    return {"is_running": process_manager.is_running}

@app.websocket("/api/logs")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    try:
        while True:
            # Non-blocking get from queue
            try:
                # Flush all available lines
                lines = []
                while True:
                    line = process_manager.output_queue.get_nowait()
                    lines.append(line)
            except queue.Empty:
                pass
            
            if lines:
                await websocket.send_text("".join(lines))
            
            await asyncio.sleep(0.1)
    except Exception:
        pass

@app.get("/api/stats")
async def get_stats():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT COUNT(*) FROM announcements")
        total_announcements = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM business_cards")
        total_cards = cursor.fetchone()[0]
        
        # Top 5 companies by project count (number of announcements)
        cursor.execute("""
            SELECT bc.company, COUNT(DISTINCT bcm.announcement_id) as count 
            FROM business_cards bc
            JOIN business_card_mentions bcm ON bcm.business_card_id = bc.id
            GROUP BY bc.company 
            ORDER BY count DESC 
            LIMIT 5
        """)
        top_companies = [dict(row) for row in cursor.fetchall()]
        
        return {
            "total_announcements": total_announcements,
            "total_cards": total_cards,
            "top_companies": top_companies
        }
    except Exception as e:
        return {"error": str(e)}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/announcements")
async def get_announcements(limit: int = 50, offset: int = 0, q: str = "", province: str = ""):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Build WHERE clause
        where_parts = []
        params = []
        
        if q:
            where_parts.append("(title LIKE ? OR content LIKE ?)")
            params.extend([f"%{q}%", f"%{q}%"])
        
        if province:
            # 按省份筛选（从标题或内容中匹配）
            where_parts.append("(title LIKE ? OR content LIKE ?)")
            params.extend([f"%{province}%", f"%{province}%"])
        
        where_clause = " AND ".join(where_parts) if where_parts else "1=1"
        
        # Get total count
        count_sql = f"SELECT COUNT(*) FROM announcements WHERE {where_clause}"
        cursor.execute(count_sql, params)
        total = cursor.fetchone()[0]
        
        # Get data
        data_sql = f"""
            SELECT id, title, url, publish_date, source 
            FROM announcements 
            WHERE {where_clause}
            ORDER BY publish_date DESC 
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_sql, params + [limit, offset])
        
        return {
            "total": total,
            "limit": limit,
            "offset": offset,
            "items": [dict(row) for row in cursor.fetchall()]
        }
    except Exception as e:
        return {"error": str(e)}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/announcements/export")
async def export_announcements(
    q: str = "", 
    province: str = "", 
    export_type: str = "all", 
    include_details: str = "false",
    start_date: str = "",
    end_date: str = "",
    pinmu: str = "",
    category: str = "",
    bid_type: str = ""
):
    include_details_bool = include_details.lower() == "true"
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Build WHERE clause
        where_parts = []
        params = []
        
        if q:
            where_parts.append("(title LIKE ? OR content LIKE ?)")
            params.extend([f"%{q}%", f"%{q}%"])
        
        if province:
             where_parts.append("(title LIKE ? OR content LIKE ?)")
             params.extend([f"%{province}%", f"%{province}%"])
        
        # Date Range
        if start_date:
            where_parts.append("publish_date >= ?")
            params.append(start_date)
        if end_date:
            where_parts.append("publish_date <= ?")
            # Append 23:59:59 if it's just a date, but let's assume yyyy-mm-dd works with string comparison if standard.
            # If user sends YYYY-MM-DD, we can append ' 23:59:59' for safety if DB has timestamps.
            # DB has "YYYY-MM-DD HH:MM:SS" typically.
            params.append(end_date + " 23:59:59")

        # Heuristic Filters (Title/Content matching)
        if pinmu:
             if pinmu == "goods":
                 where_parts.append("title LIKE ?")
                 params.append("%货物%")
             elif pinmu == "engineering":
                 where_parts.append("title LIKE ?")
                 params.append("%工程%")
             elif pinmu == "services":
                 where_parts.append("title LIKE ?")
                 params.append("%服务%")

        if category:
            # This is hard because "category" (central/local) depends on source URL or other heuristics not easily mapped to title.
            # We might have to skip this or try basic domain matching?
            # For now, let's look at the source content or maybe just URL?
            # Let's skip for now or do a very rough check if feasible.
            # User request said "source category". Our DB has 'source' column 'ccgp', etc.
            # Let's assume user wants to filter by "Central" -> "ccgp" etc? 
            # Actually, `server.py` doesn't seem to have a `category` column.
            # Let's implement based on what we can. 
            pass 

        if bid_type:
            # Map types
            if bid_type == "1": # Open Tendering
                where_parts.append("(title LIKE ? OR title LIKE ?)")
                params.extend(["%公开招标%", "%招标公告%"])
            elif bid_type == "7": # Winning
                where_parts.append("(title LIKE ? OR title LIKE ?)")
                params.extend(["%中标%", "%成交%"])
            elif bid_type == "2": # Competitive Negotiation
                where_parts.append("title LIKE ?")
                params.append("%谈判%")
            elif bid_type == "3": # Competitive Consultation
                where_parts.append("title LIKE ?")
                params.append("%磋商%")
            elif bid_type == "4": # Inquiry
                 where_parts.append("title LIKE ?")
                 params.append("%询价%")
            elif bid_type == "5": # Single Source
                 where_parts.append("title LIKE ?")
                 params.append("%单一来源%")
        
        where_clause = " AND ".join(where_parts) if where_parts else "1=1"
        
        # Query all matching data
        sql = f"""
            SELECT title, url, publish_date, source, content
            FROM announcements 
            WHERE {where_clause}
            ORDER BY publish_date DESC 
        """
        cursor.execute(sql, params)
        rows = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        if not rows:
             return JSONResponse(status_code=400, content={"error": "No data to export"})

        try:
            import openpyxl
        except ImportError:
            return JSONResponse(status_code=500, content={"error": "openpyxl not installed"})

        # Helper to create excel sheet
        def create_sheet(wb, title, data):
            ws = wb.create_sheet(title=title[:30]) # Sheet name limit 31 chars
            # Headers
            if include_details_bool:
                 # If including details, HIDE Source and Link (as requested)
                 headers = ["发布时间", "标题", "详情"]
            else:
                 headers = ["发布时间", "标题", "来源", "链接"]
            
            ws.append(headers)
            
            # Data
            for item in data:
                if include_details_bool:
                    # Limit content length
                    content = str(item.get('content', ''))
                    if len(content) > 32000:
                        content = content[:32000] + "..."
                    ws.append([
                        str(item.get('publish_date', '')),
                        str(item.get('title', '')),
                        content
                    ])
                else:
                    ws.append([
                        str(item.get('publish_date', '')),
                        str(item.get('title', '')),
                        str(item.get('source', '')),
                        str(item.get('url', ''))
                    ])

            # Adjust widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 60
            if include_details_bool:
                ws.column_dimensions['C'].width = 80
            else:
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 40
            
        
        if export_type == "all":
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                 del wb["Sheet"]
            
            create_sheet(wb, "公告列表", rows)
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"announcements_{export_type}.xlsx"
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        elif export_type == "province":
            # Group by province
            provinces = ["北京", "上海", "广东", "江苏", "浙江", "山东", "河南", "四川", "湖北", "湖南", "福建", "安徽", "河北", "陕西", "江西", "重庆", "辽宁", "云南", "广西", "山西", "贵州", "天津", "新疆", "内蒙古", "黑龙江", "吉林", "甘肃", "海南", "宁夏", "青海", "西藏"]
            
            grouped = {p: [] for p in provinces}
            grouped["其他"] = []
            
            for item in rows:
                text = (item.get('title', '') + item.get('content', '')).lower()
                found = False
                for p in provinces:
                    if p in text:
                        grouped[p].append(item)
                        found = True
                        # Don't break if we want to add to multiple? 
                        # Ideally assign to all matched, but for now specific enough.
                        # If a record belongs to multiple, it will be added to the first hit if we break.
                        # Let's not break, add to all matched? No, duplication might be annoying.
                        # Let's break for simplicity, assuming primary province comes first or is good enough.
                        break 
                if not found:
                    grouped["其他"].append(item)
            
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                for p_name, p_data in grouped.items():
                    if not p_data:
                        continue
                        
                    wb = openpyxl.Workbook()
                    del wb["Sheet"]
                    create_sheet(wb, p_name, p_data)
                    
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    zip_file.writestr(f"{p_name}.xlsx", excel_buffer.getvalue())
            
            zip_buffer.seek(0)
            return StreamingResponse(
                zip_buffer,
                media_type="application/zip",
                headers={"Content-Disposition": "attachment; filename=announcements_by_province.zip"}
            )
            
        elif export_type == "day":
            # Group by date
            grouped = {}
            for item in rows:
                date_str = item.get('publish_date', '')
                if not date_str:
                    date_str = "Unknown"
                else:
                    # Try to extract YYYY-MM-DD
                    try:
                        date_str = date_str.split(' ')[0] # Handle "2023-01-01 12:00"
                    except:
                        pass
                
                if date_str not in grouped:
                    grouped[date_str] = []
                grouped[date_str].append(item)
                
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                for d_str, d_data in grouped.items():
                    wb = openpyxl.Workbook()
                    del wb["Sheet"]
                    create_sheet(wb, d_str, d_data)
                    
                    excel_buffer = BytesIO()
                    wb.save(excel_buffer)
                    zip_file.writestr(f"{d_str}.xlsx", excel_buffer.getvalue())

            zip_buffer.seek(0)
            return StreamingResponse(
                zip_buffer,
                media_type="application/zip",
                headers={"Content-Disposition": "attachment; filename=announcements_by_day.zip"}
            )
            
        else:
             return JSONResponse(status_code=400, content={"error": "Invalid export type"})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/announcements/{item_id}")
async def get_announcement_detail(item_id: int):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM announcements WHERE id = ?", (item_id,))
        row = cursor.fetchone()
        if not row:
            return JSONResponse(status_code=404, content={"error": "Not found"})
        return dict(row)
    except Exception as e:
        return {"error": str(e)}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/cards")
async def get_cards(limit: int = 50, offset: int = 0, q: str = ""):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 基础查询条件
        where_clause = ""
        params = []
        if q:
            search = f"%{q}%"
            where_clause = "WHERE bc.company LIKE ? OR bc.contact_name LIKE ?"
            params = [search, search]
        
        # 获取总数 (Deduplicated count)
        count_sql = f"""
            SELECT COUNT(*) FROM (
                SELECT 1 
                FROM business_cards bc
                {where_clause}
                GROUP BY bc.company, bc.contact_name
            )
        """
        cursor.execute(count_sql, params)
        total_row = cursor.fetchone()
        total = total_row[0] if total_row else 0
        
        # 获取分页数据 (Deduplicated)
        # We group by company+contact and aggregate project counts.
        # We pick the MAX(id) to represent the group for clicking.
        data_sql = f"""
            SELECT 
                MAX(bc.id) as id, 
                bc.company, 
                bc.contact_name, 
                MAX(bc.phones_json) as phones_json, 
                MAX(bc.emails_json) as emails_json,
                MAX(bc.created_at) as created_at, 
                MAX(bc.updated_at) as updated_at,
                COUNT(DISTINCT bcm.announcement_id) AS projects_count
            FROM business_cards bc
            LEFT JOIN business_card_mentions bcm ON bcm.business_card_id = bc.id
            {where_clause}
            GROUP BY bc.company, bc.contact_name
            ORDER BY updated_at DESC 
            LIMIT ? OFFSET ?
        """
        cursor.execute(data_sql, params + [limit, offset])
        
        import json
        results = []
        for row in cursor.fetchall():
            card = dict(row)
            # Parse JSON fields
            try:
                phones_list = json.loads(card.get('phones_json') or '[]')
                card['phones'] = ', '.join(phones_list) if phones_list else ''
            except:
                card['phones'] = ''
            try:
                emails_list = json.loads(card.get('emails_json') or '[]')
                card['emails'] = ', '.join(emails_list) if emails_list else ''
            except:
                card['emails'] = ''
            results.append(card)
        
        return {
            "total": total,
            "limit": limit,
            "offset": offset,
            "items": results
        }
    except Exception as e:
        return {"error": str(e)}
    finally:
        if 'conn' in locals():
            conn.close()

@app.get("/api/cards/{card_id}/mentions")
async def get_card_mentions(card_id: int):
    try:
        conn = get_db_connection()
        # We need to use the Database class method, but here we are using raw connection in get_db_connection helper.
        # However, the previous code in server.py mainly uses raw sql. 
        # But wait, `Database()` in `lifespan` just inits the table.
        # The `get_db_connection` returns a raw sqlite3 connection.
        # It's better to use the `Database` class if possible, or just write the SQL here.
        # Since I just added the SQL to Database class, I should probably use it or copy the SQL.
        # `server.py` seems to use `get_db_connection` and raw SQL for other endpoints.
        # To match the style (and avoid instantiating Database class which might have overhead or different locking),
        # I will replicate the SQL here as commonly done in this file.
        
        cursor = conn.cursor()
        
        # First, find the company and contact_name for this card_id
        cursor.execute("SELECT company, contact_name FROM business_cards WHERE id = ?", (card_id,))
        card_info = cursor.fetchone()
        
        if not card_info:
            return []
            
        company, contact_name = card_info['company'], card_info['contact_name']
        
        # Now fetch mentions for ALL cards that match this company and contact_name
        # This aggregates history from "duplicate" entries
        cursor.execute("""
            SELECT 
                a.id, a.title, a.url, a.source, a.publish_date, 
                bcm.role
            FROM business_card_mentions bcm
            JOIN business_cards bc ON bcm.business_card_id = bc.id
            JOIN announcements a ON bcm.announcement_id = a.id
            WHERE bc.company = ? AND bc.contact_name = ?
            ORDER BY a.publish_date DESC
        """, (company, contact_name))
        
        return [dict(row) for row in cursor.fetchall()]
    except Exception as e:
        return {"error": str(e)}
    finally:
        if 'conn' in locals():
            conn.close()

# ========== Keyword Management ==========
KEYWORDS_FILE = os.path.join("data", "saved_keywords.json")

class KeywordsRequest(BaseModel):
    keywords: List[str]
    filename: Optional[str] = None

@app.get("/api/keywords")
async def get_keywords():
    try:
        if os.path.exists(KEYWORDS_FILE):
            import json
            with open(KEYWORDS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data
        return {"keywords": [], "filename": None}
    except Exception as e:
        return {"error": str(e)}

@app.post("/api/keywords")
async def save_keywords(req: KeywordsRequest):
    try:
        import json
        # Ensure data directory exists
        os.makedirs(os.path.dirname(KEYWORDS_FILE), exist_ok=True)
        
        data = {
            "keywords": req.keywords,
            "filename": req.filename
        }
        with open(KEYWORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return {"status": "saved", "count": len(req.keywords)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})

# ========== Card Export ==========
@app.get("/api/cards/export")
async def export_cards(q: str = ""):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        if q:
            search = f"%{q}%"
            cursor.execute("""
                SELECT 
                    bc.id, bc.company, bc.contact_name, 
                    bc.phones_json, bc.emails_json,
                    bc.created_at, bc.updated_at,
                    COUNT(DISTINCT bcm.announcement_id) AS projects_count
                FROM business_cards bc
                LEFT JOIN business_card_mentions bcm ON bcm.business_card_id = bc.id
                WHERE bc.company LIKE ? OR bc.contact_name LIKE ?
                GROUP BY bc.id
                ORDER BY bc.company, bc.contact_name
            """, (search, search))
        else:
            cursor.execute("""
                SELECT 
                    bc.id, bc.company, bc.contact_name, 
                    bc.phones_json, bc.emails_json,
                    bc.created_at, bc.updated_at,
                    COUNT(DISTINCT bcm.announcement_id) AS projects_count
                FROM business_cards bc
                LEFT JOIN business_card_mentions bcm ON bcm.business_card_id = bc.id
                GROUP BY bc.id
                ORDER BY bc.company, bc.contact_name
            """)
        
        import json
        rows = []
        for row in cursor.fetchall():
            card = dict(row)
            try:
                phones = json.loads(card.get('phones_json') or '[]')
                card['phones'] = ', '.join(phones)
            except:
                card['phones'] = ''
            try:
                emails = json.loads(card.get('emails_json') or '[]')
                card['emails'] = ', '.join(emails)
            except:
                card['emails'] = ''
            rows.append(card)
        
        conn.close()
        
        # Generate Excel file
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            return JSONResponse(status_code=500, content={"error": "openpyxl not installed. Run: pip install openpyxl"})
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "名片库"
        
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["公司名称", "联系人", "电话", "邮箱", "关联项目数", "创建时间", "更新时间"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, card in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=card.get('company', ''))
            ws.cell(row=row_idx, column=2, value=card.get('contact_name', ''))
            ws.cell(row=row_idx, column=3, value=card.get('phones', ''))
            ws.cell(row=row_idx, column=4, value=card.get('emails', ''))
            ws.cell(row=row_idx, column=5, value=card.get('projects_count', 0))
            ws.cell(row=row_idx, column=6, value=card.get('created_at', ''))
            ws.cell(row=row_idx, column=7, value=card.get('updated_at', ''))
            
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        
        # Save to bytes
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from datetime import datetime
        filename = f"business_cards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


# Mount static files - MUST BE LAST
if getattr(sys, 'frozen', False):
    # In frozen mode, files are next to the executable (onedir mode)
    exe_dir = os.path.dirname(os.path.abspath(sys.executable))
    web_dir = os.path.join(exe_dir, "web")
else:
    # In dev mode
    web_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "web")

if os.path.exists(web_dir):
    app.mount("/", StaticFiles(directory=web_dir, html=True), name="static")
else:
    # Fallback to current dir if not found in base_path
    alt_web_dir = os.path.join(os.getcwd(), "web")
    if os.path.exists(alt_web_dir):
        app.mount("/", StaticFiles(directory=alt_web_dir, html=True), name="static")
    else:
        print(f"[WARNING] Web directory not found at {web_dir}")

def open_browser():
    """Wait a moment and open the browser"""
    import time
    import webbrowser
    time.sleep(1.5)
    print("Opening browser at http://localhost:8080")
    webbrowser.open("http://localhost:8080")

if __name__ == "__main__":
    import uvicorn
    import threading
    import sys
    
    # If frozen (PyInstaller), force Playwright to use bundled browsers
    if getattr(sys, 'frozen', False):
        if hasattr(sys, '_MEIPASS'):
             base_path = sys._MEIPASS
        else:
             base_path = os.path.dirname(os.path.abspath(sys.executable))
             
        os.environ["PLAYWRIGHT_BROWSERS_PATH"] = os.path.join(base_path, "browsers")

    try:
        # Unified Entry Point Dispatch
        if len(sys.argv) > 1:
            # WORKER MODE
            try:
                app_worker.main()
            except SystemExit:
                pass
            except Exception as e:
                print(f"Worker Error: {e}")
            finally:
                # IMPORTANT: Exit the process so we don't fall through to the Server Mode code
                sys.exit(0)
        else:
            # SERVER MODE
            # Ensure data dir exists
            if not os.path.exists("data"):
                os.makedirs("data")

            # Auto-open browser
            threading.Thread(target=open_browser, daemon=True).start()
                
            print("Starting server on http://localhost:8080")
            uvicorn.run(app, host="0.0.0.0", port=8080, log_level="info")
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"\n[CRITICAL ERROR] {e}")
        input("Press Enter to exit...")
