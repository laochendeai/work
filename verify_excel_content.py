
import sqlite3
import os
import openpyxl
from io import BytesIO

def get_db_connection():
    db_path = os.path.join("data", "gp.db")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn

def create_sheet(wb, title, data):
    ws = wb.create_sheet(title=title[:30]) # Sheet name limit 31 chars
    # Headers
    headers = ["发布时间", "标题", "来源", "链接"]
    ws.append(headers)
    # Data
    for item in data:
        ws.append([
            item.get('publish_date', ''),
            item.get('title', ''),
            item.get('source', ''),
            item.get('url', '')
        ])
    # Adjust widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30

def verify_excel():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT title, url, publish_date, source, content FROM announcements LIMIT 5")
        rows = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        if not rows:
            print("No rows in DB")
            return

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
             del wb["Sheet"]
        
        create_sheet(wb, "Test", rows)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Read back
        wb2 = openpyxl.load_workbook(output)
        ws2 = wb2.active
        
        print(f"Sheet name: {ws2.title}")
        header = [cell.value for cell in ws2[1]]
        print(f"Header: {header}")
        
        first_row = [cell.value for cell in ws2[2]]
        print(f"First row: {first_row}")
        
        if "来源" not in header:
            print("FAIL: Header '来源' missing")
        else:
            idx = header.index("来源")
            print(f"Source column index: {idx}")
            print(f"Source value in first row: {first_row[idx]}")

    except Exception as e:
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verify_excel()
